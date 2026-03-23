import logging
import os
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from typing import Optional

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ----------------------------------------------
# CONFIG
# ----------------------------------------------

INPUT_FILE = "input.xlsx"
OUTPUT_FILE = "output.xlsx"

MAX_WORKERS = 5
MAX_RETRIES = 3
REQUEST_TIMEOUT = 20

IBASE_URL = "https://pcsupport.lenovo.com/us/en/api/v4/upsell/redport/getIbaseInfo"

HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Content-Type": "application/json"
}

STATUS_SUCCESS = "SUCCESS"
STATUS_FAILED = "FAILED"
STATUS_INVALID = "INVALID"

# ----------------------------------------------
# LOGGING
# ----------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(threadName)s - %(message)s",
    handlers=[logging.FileHandler("app.log"), logging.StreamHandler()],
)

logger = logging.getLogger(__name__)

# ----------------------------------------------
# MODEL
# ----------------------------------------------

@dataclass
class LaptopRecord:
    serial_number: str
    product_name: str = ""
    machine_type: str = ""
    cpu: str = ""
    ram: str = ""
    storage: str = ""
    display: str = ""
    status: str = STATUS_FAILED
    error: str = ""

# ----------------------------------------------
# VALIDATION
# ----------------------------------------------

def validate_serial(serial):
    if not serial or not serial.strip():
        return False, "Empty serial"
    if not re.fullmatch(r"[A-Za-z0-9]{5,20}", serial):
        return False, "Invalid format"
    return True, ""

# ----------------------------------------------
# RETRY REQUEST
# ----------------------------------------------

def request_with_retry(url, payload):
    for attempt in range(MAX_RETRIES):
        try:
            res = requests.post(url, json=payload, headers=HEADERS, timeout=REQUEST_TIMEOUT)
            res.raise_for_status()
            return res.json()
        except Exception as e:
            logger.warning(f"Retry {attempt+1}: {e}")
            time.sleep(2 ** attempt)
    return None

# ----------------------------------------------
# GET BASIC INFO (FIXED)
# ----------------------------------------------

def get_basic_info(serial):
    payload = {
        "serialNumber": serial,
        "country": "us",
        "language": "en"
    }

    data = request_with_retry(IBASE_URL, payload)

    if not data:
        return None

    try:
        machine_info = data.get("data", {}).get("machineInfo", {})

        return {
            "productName": machine_info.get("productName", ""),
            "machineType": machine_info.get("type", ""),
            "specHTML": machine_info.get("specification", "")
        }
    except:
        return None

# ----------------------------------------------
# HTML SPEC PARSER (NEW CORE LOGIC)
# ----------------------------------------------

def extract_specs_from_html(html):
    specs = {"cpu": "N/A", "ram": "N/A", "storage": "N/A", "display": "N/A"}

    if not html:
        return specs

    soup = BeautifulSoup(html, "lxml")
    rows = soup.find_all("tr")

    for row in rows:
        cols = row.find_all("td")
        if len(cols) != 2:
            continue

        key = cols[0].get_text(strip=True).lower()
        value = cols[1].get_text(strip=True)

        if "processor" in key:
            specs["cpu"] = value
        elif "memory" in key:
            specs["ram"] = value
        elif "hard drive" in key or "storage" in key:
            specs["storage"] = value
        elif "monitor" in key or "display" in key:
            specs["display"] = value

    return specs

# ----------------------------------------------
# PROCESS SERIAL
# ----------------------------------------------

def process_serial(serial):
    record = LaptopRecord(serial_number=serial)

    valid, reason = validate_serial(serial)
    if not valid:
        record.status = STATUS_INVALID
        record.error = reason
        return record

    logger.info(f"Processing {serial}")

    basic = get_basic_info(serial)

    if not basic:
        record.status = STATUS_FAILED
        record.error = "API failed"
        return record

    record.product_name = basic["productName"]
    record.machine_type = basic["machineType"]

    specs = extract_specs_from_html(basic["specHTML"])

    record.cpu = specs["cpu"]
    record.ram = specs["ram"]
    record.storage = specs["storage"]
    record.display = specs["display"]

    record.status = STATUS_SUCCESS

    return record

# ----------------------------------------------
# EXCEL HELPERS
# ----------------------------------------------

def read_input():
    df = pd.read_excel(INPUT_FILE)
    return df["SerialNumber"].fillna("").tolist()

def write_output(records):
    rows = [r.__dict__ for r in records]
    df = pd.DataFrame(rows)
    df.to_excel(OUTPUT_FILE, index=False)

# ----------------------------------------------
# MAIN
# ----------------------------------------------

def main():
    serials = read_input()

    results = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = [executor.submit(process_serial, s) for s in serials]

        for f in as_completed(futures):
            results.append(f.result())

    write_output(results)

    print("\nDone. Check output.xlsx")

# ----------------------------------------------

if __name__ == "__main__":
    main()